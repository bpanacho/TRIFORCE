
# Importação de bibliotecas necessárias
import cv2  # OpenCV para processamento de imagem
import pickle  # Para serialização de objetos Python
import extrairGabarito as exG  # Módulo personalizado para extrair gabaritos
import keyboard  # Biblioteca para captura de eventos do teclado
import openpyxl  # Biblioteca para manipulação de arquivos Excel
from flask import Flask, render_template, Response  # Flask para criar um servidor web

# Inicialização da aplicação Flask
app = Flask(__name__)

# Inicialização de variáveis globais
camera = cv2.VideoCapture(0)  # Inicialização da câmera
respostasCorretas = None  # Inicialização da variável global para respostas corretas

# Definição das rotas da aplicação web
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/termos')
def termos():
    return render_template('termos.html')

@app.route('/videos')
def videos():
    return render_template('videos.html')

global pontuacao
pontuacao = 0


# Função para processar o vídeo da webcam
def processar_video():
    global respostasCorretas
      # Utilização da variável global

    # Inicialização da câmera
    video = cv2.VideoCapture(0)

    # Verificação da abertura da câmera
    if not video.isOpened():
        print("Não foi possível abrir a webcam. Verifique se está conectada corretamente.")
        return

    # Leitura dos campos e respostas de arquivos pickle
    campos = []
    resp = []
    with open('C:\\Users\\pedro\\Downloads\\formulario-animado-main (2)\\formulario-animado-main\\campos.pkl', 'rb') as arquivo_campos:
        campos = pickle.load(arquivo_campos)

    with open('C:\\Users\\pedro\\Downloads\\formulario-animado-main (2)\\formulario-animado-main\\resp.pkl', 'rb') as arquivo_respostas:
        resp = pickle.load(arquivo_respostas)

    # Inicialização de variáveis para controle do fluxo
    pausado = False
    imprimir_respostas = False

    # Lista para armazenar respostas corretas
    respostasCorretas = []

    # Loop principal de captura de vídeo
    while True:
        _, imagem = video.read()  # Captura do frame da webcam

        # Verificação de erro na captura do frame
        if not _:
            print("Falha ao capturar frame da webcam.")
            break

        # Redimensionamento da imagem
        imagem = cv2.resize(imagem, (600, 700))
        
        # Extração do gabarito da imagem
        gabarito, bbox = exG.extrairMaiorCtn(imagem)

        if gabarito is not None and bbox is not None:
            # Pré-processamento do gabarito
            imgGray = cv2.cvtColor(gabarito, cv2.COLOR_BGR2GRAY)
            ret, imgTh = cv2.threshold(imgGray, 70, 255, cv2.THRESH_BINARY_INV)
            cv2.rectangle(imagem, (bbox[0], bbox[1]), (bbox[0] + bbox[2], bbox[1] + bbox[3]), (255, 165, 0), 3)

            respostas = []

            # Iteração sobre os campos identificados
            for id, vg in enumerate(campos):
                x, y, w, h = int(vg[0]), int(vg[1]), int(vg[2]), int(vg[3])
                cv2.rectangle(gabarito, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.rectangle(imgTh, (x, y), (x + w, y + h), (255, 255, 255), 1)
                campo = imgTh[y:y + h, x:x + w]

                percentual_preto = (cv2.countNonZero(campo) / (h * w)) * 100

                if percentual_preto >= 10:
                    cv2.rectangle(gabarito, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    respostas.append(resp[id])

            # Exibição das imagens e resultados
            cv2.imshow('Gabarito', gabarito)
            
            # Conversão da imagem para formato JPEG e envio para a interface web
            _, buffer_imagem = cv2.imencode('.jpg', imagem)
            frame_bytes_imagem = buffer_imagem.tobytes()

            yield (b'--frame\r\n'
                b'Content-Type: image/jpeg\r\n'
                b'Content-Length: ' + f"{len(frame_bytes_imagem)}".encode() + b'\r\n\r\n' + frame_bytes_imagem + b'\r\n')
            
            # Verificação da tecla 'p' para pausar/despausar
            if keyboard.is_pressed('p'):
                pausado = not pausado
                print("Pausado" if pausado else "Despausado")

            # Verificação da tecla 's' para imprimir respostas
            if keyboard.is_pressed('s') and not pausado and not imprimir_respostas:
                imprimir_respostas = True
                print("Respostas: ", respostas)
                respostasCorretas = respostas

        # Verificação da tecla 'q' para encerrar o loop
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Liberação da webcam e fechamento das janelas
    video.release()
    cv2.destroyAllWindows()
    
    # Criação de um arquivo Excel para armazenar pontuações
    workbook = openpyxl.Workbook()
    # Configuração do codec de vídeo

    fourcc = cv2.VideoWriter_fourcc(*'H264')
    print(f"Suporta o codec H264? {cv2.VideoWriter_fourcc(*'H264') != 0}")

    # Inicialização da variável de pausa
    paused = False

    # Reabertura da câmera
    video = cv2.VideoCapture(0)

    # Verificação da abertura da câmera
    if not video.isOpened():
        print("Não foi possível abrir a webcam. Verifique se está conectada corretamente.")
        exit()

    # Leitura dos campos e respostas de arquivos pickle
    campos = []
    resp = []
    with open('C:\\Users\\pedro\\Downloads\\formulario-animado-main (2)\\formulario-animado-main\\campos.pkl', 'rb') as arquivo_campos:
        campos = pickle.load(arquivo_campos)

    with open('C:\\Users\\pedro\\Downloads\\formulario-animado-main (2)\\formulario-animado-main\\resp.pkl', 'rb') as arquivo_respostas:
        resp = pickle.load(arquivo_respostas)

    # Seleção da planilha ativa no arquivo Excel
    sheet = workbook.active

    # Definição dos cabeçalhos na primeira linha da planilha
    sheet['A1'] = 'Prova'
    sheet['B1'] = 'Acertos'
    sheet['C1'] = 'Erros'
    sheet['D1'] = 'Pontuação'

    # Inicialização da variável de prova fora do loop de captura
    prova = 1

    # Loop para captura contínua de vídeo
    while True:
        key = cv2.waitKey(1) & 0xFF

        if key == ord('q'):
            break

        if not paused:
            _, imagem = video.read()
            imagem = cv2.resize(imagem, (600, 700))
            gabarito, bbox = exG.extrairMaiorCtn(imagem)

            if gabarito is not None and bbox is not None:
                imgGray = cv2.cvtColor(gabarito, cv2.COLOR_BGR2GRAY)
                ret, imgTh = cv2.threshold(imgGray, 70, 255, cv2.THRESH_BINARY_INV)    
                cv2.rectangle(imagem, (bbox[0], bbox[1]), (bbox[0] + bbox[2], bbox[1] + bbox[3]), (255, 165, 0), 3)

                respostas = []
                for id, vg in enumerate(campos):
                    x, y, w, h = int(vg[0]), int(vg[1]), int(vg[2]), int(vg[3])
                    cv2.rectangle(gabarito, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    cv2.rectangle(imgTh, (x, y), (x + w, y + h), (255, 255, 255), 1)
                    campo = imgTh[y:y + h, x:x + w]

                    percentual_preto = (cv2.countNonZero(campo) / (h * w)) * 100
                    if percentual_preto >= 10:
                        cv2.rectangle(gabarito, (x, y), (x + w, y + h), (255, 0, 0), 2)
                        respostas.append(resp[id])

                erros = 0
                acertos = 0
                if len(respostas) == len(respostasCorretas):
                    for num, res in enumerate(respostas):
                        if res == respostasCorretas[num]:
                            acertos += 1
                        else:
                            erros += 1

                    pontuacao = int(acertos * 1)
                cv2.putText(imagem, 'Pontuacao:', (30, 140), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 0, 0), 3)
                cv2.putText(imagem, 'Acertos:', (30, 180), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 0, 0), 3)
                cv2.putText(imagem, 'Erros:', (30, 220), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 0, 0), 3)

                cv2.putText(imagem, f'{pontuacao}', (250, 140), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (255, 0, 0), 3)
                cv2.putText(imagem, f'{acertos}', (190, 180), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 158, 0), 3)
                cv2.putText(imagem, f'{erros}', (180, 220), cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 0, 255), 3)                            

            if paused:
                cv2.putText(imagem, 'Vídeo Pausado', (30, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

            # Envio da imagem para a interface web
            _, buffer_imagem = cv2.imencode('.jpg', imagem)
            frame_bytes_imagem = buffer_imagem.tobytes()

            yield (b'--frame\r\n'
                b'Content-Type: image/jpeg\r\n'
                b'Content-Length: ' + f"{len(frame_bytes_imagem)}".encode() + b'\r\n\r\n' + frame_bytes_imagem + b'\r\n')

        if key != 255:
            # Atualização da planilha Excel após pressionar qualquer tecla (exceto 'q')
            prova += 1
            sheet.cell(row=prova, column=1, value=prova - 1)
            sheet.cell(row=prova, column=2, value=acertos)
            sheet.cell(row=prova, column=3, value=erros)
            sheet.cell(row=prova, column=4, value=pontuacao)

        # Salvamento do arquivo Excel
        workbook.save("pontuacao.xlsx")

    # Liberação da webcam e fechamento das janelas
    video.release()
    cv2.destroyAllWindows()

# Rota para o feed de vídeo
@app.route('/video_feed')   
def video_feed():
    return Response(processar_video(), mimetype='multipart/x-mixed-replace; boundary=frame')

# Execução da aplicação Flask
if __name__ == '__main__':
    app.run(debug=True)

