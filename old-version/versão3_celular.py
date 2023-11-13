
import cv2
import pickle
import extrairGabarito as exG
import keyboard
import openpyxl
from flask import Flask, render_template, Response

# app = Flask(__name__, template_folder='templates', static_folder='static')
app = Flask(__name__)

# PROFESSOR
# CADASTRAR GABARITO
#camera = cv2.VideoCapture('http://192.168.0.119:4747/video')

respostasCorretas = None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/termos')
def termos():
    return render_template('termos.html')

@app.route('/videos')
def videos():
    return render_template('videos.html')
            
def processar_video():
    global respostasCorretas
    
    video = cv2.VideoCapture('http://192.168.0.119:4747/video')

    if not video.isOpened():
        print("Não foi possível abrir a webcam. Verifique se está conectada corretamente.")
        return

    campos = []
    resp = []
    with open('C:\\Users\\pedro\\Downloads\\formulario-animado-main (2)\\formulario-animado-main\\campos.pkl', 'rb') as arquivo_campos:
        campos = pickle.load(arquivo_campos)

# Para abrir o arquivo resp.pkl
    with open('formulario-animado-main/resp.pkl', 'rb') as arquivo_respostas:
        resp = pickle.load(arquivo_respostas)

    pausado = False
    imprimir_respostas = False
    respostasCorretas = []  # Adicione uma lista para as respostas corretas

    while True:
        _, imagem = video.read()

        if not _:
            print("Falha ao capturar frame da webcam.")
            break

        imagem = cv2.resize(imagem, (600, 700))
        gabarito, bbox = exG.extrairMaiorCtn(imagem)

        if gabarito is not None and bbox is not None:
            imgGray = cv2.cvtColor(gabarito, cv2.COLOR_BGR2GRAY)
            ret, imgTh = cv2.threshold(imgGray, 70, 255, cv2.THRESH_BINARY_INV)
            cv2.rectangle(imagem, (bbox[0], bbox[1]), (bbox[0] + bbox[2], bbox[1] + bbox[3]), (0, 255, 0), 3)

            respostas = []
            for id, vg in enumerate(campos):
                x, y, w, h = int(vg[0]), int(vg[1]), int(vg[2]), int(vg[3])
                cv2.rectangle(gabarito, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.rectangle(imgTh, (x, y), (x + w, y + h), (255, 255, 255), 1)
                campo = imgTh[y:y + h, x:x + w]

                percentual_preto = (cv2.countNonZero(campo) / (h * w)) * 100
                if percentual_preto >= 10:
                    cv2.rectangle(gabarito, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    respostas.append(resp[id])

            # Exibir imagens e resultados
            # cv2.imshow('img', imagem)
            cv2.imshow('Gabarito', gabarito)
            # cv2.imshow('IMG TH', imgTh)
            
            _, buffer_imagem = cv2.imencode('.jpg', imagem)
            frame_bytes_imagem = buffer_imagem.tobytes()

            yield (b'--frame\r\n'
                b'Content-Type: image/jpeg\r\n'
                b'Content-Length: ' + f"{len(frame_bytes_imagem)}".encode() + b'\r\n\r\n' + frame_bytes_imagem + b'\r\n')
            
            # Verificar se a tecla 'p' foi pressionada para pausar/despausar
            if keyboard.is_pressed('p'):
                pausado = not pausado
                print("Pausado" if pausado else "Despausado")

            # Verificar se a tecla 's' foi pressionada para imprimir respostas
            if keyboard.is_pressed('s') and not pausado and not imprimir_respostas:
                imprimir_respostas = True
                print("Respostas: ", respostas)
                respostasCorretas = respostas  # Defina as respostas corretas

        # Liberar a webcam e fechar todas as janelas quando 'q' for pressionado
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Liberar a webcam e fechar todas as janelas
    # video.release()
    cv2.destroyAllWindows()
    
    workbook = openpyxl.Workbook()

    fourcc = cv2.VideoWriter_fourcc(*'H264')
    print(f"Suporta o codec H264? {cv2.VideoWriter_fourcc(*'H264') != 0}")

    # Inicializa a variável de pausa
    paused = False

    # video = cv2.VideoCapture('http://192.168.2.107:4747/video')

    if not video.isOpened():
        print("Não foi possível abrir a webcam. Verifique se está conectada corretamente.")
        exit()


        campos = []
        resp = []
    with open('C:\\Users\pedro\\Downloads\\formulario-animado-main (2)\\formulario-animado-main\\campos.pkl', 'rb') as arquivo_campos:
        campos = pickle.load(arquivo_campos)

# Para abrir o arquivo resp.pkl
    with open('formulario-animado-main/resp.pkl', 'rb') as arquivo_respostas:
     resp = pickle.load(arquivo_respostas)

    # Selecione a planilha ativa
    sheet = workbook.active

    # Defina os cabeçalhos na primeira linha
    sheet['A1'] = 'Prova'
    sheet['B1'] = 'Acertos'
    sheet['C1'] = 'Erros'
    sheet['D1'] = 'Pontuação'

    # Inicialize a variável prova fora do loop de captura
    prova = 1


    while True:
        key = cv2.waitKey(1) & 0xFF

        if key == ord('q'):
            break
        
        erros = 0
        acertos = 0
        total_questoes = len(respostasCorretas)

        if total_questoes > 0:
            for num, res in enumerate(respostas):
                if num < total_questoes and res == respostasCorretas[num]:
                    acertos += 1
            else:
                erros += 1  # Conta respostas erradas e em branco

            pontuacao = int(acertos * 1)  # Por exemplo, +1 ponto por resposta correta
        else:
            pontuacao = 0

        if not paused:
            _, imagem = video.read()
            imagem = cv2.resize(imagem, (600, 700))
            gabarito, bbox = exG.extrairMaiorCtn(imagem)

            if gabarito is not None and bbox is not None:
                imgGray = cv2.cvtColor(gabarito, cv2.COLOR_BGR2GRAY)
                ret, imgTh = cv2.threshold(imgGray, 70, 255, cv2.THRESH_BINARY_INV)
                cv2.rectangle(imagem, (bbox[0], bbox[1]), (bbox[0] + bbox[2], bbox[1] + bbox[3]), (0, 255, 0), 3)

                respostas = []
                for id, vg in enumerate(campos):
                    x, y, w, h = int(vg[0]), int(vg[1]), int(vg[2]), int(vg[3])
                    cv2.rectangle(gabarito, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    cv2.rectangle(imgTh, (x, y), (x + w, y + h), (255, 255, 255), 1)
                    campo = imgTh[y:y + h, x:x + w]

                    percentual_preto = (cv2.countNonZero(campo) / (h * w)) * 100
                    if percentual_preto >= 10:
                        cv2.rectangle(gabarito, (x, y), (x + w, y + h), (255, 0, 0), 2)
                        respostas.append(resp[id])

                erros = 0
                acertos = 0
                if len(respostas) == len(respostasCorretas):
                    for num, res in enumerate(respostas):
                        if res == respostasCorretas[num]:
                            acertos += 1
                        else:
                            erros += 1

                    pontuacao = int(acertos * 1)
                    cv2.putText(imagem, f'ACERTOS: {acertos}, PONTOS: {pontuacao}', (30, 140), cv2.FONT_HERSHEY_SIMPLEX, 1.2,
                                (0, 0, 255), 3)

            # Adiciona um texto para indicar quando o vídeo está pausado
            if paused:
                cv2.putText(imagem, 'Vídeo Pausado', (30, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

            # cv2.imshow('img', imagem)
            cv2.imshow('Gabarito do Aluno', gabarito)
            # cv2.imshow('IMG TH', imgTh)
            
            _, buffer_imagem = cv2.imencode('.jpg', imagem)
            frame_bytes_imagem = buffer_imagem.tobytes()

            yield (b'--frame\r\n'
                b'Content-Type: image/jpeg\r\n'
                b'Content-Length: ' + f"{len(frame_bytes_imagem)}".encode() + b'\r\n\r\n' + frame_bytes_imagem + b'\r\n')

        if key != 255:  # Qualquer tecla, exceto 'q'
        # Salve o número de questões acertadas na planilha
            prova += 1
        sheet.cell(row=prova, column=1, value=prova - 1)
        sheet.cell(row=prova, column=2, value=acertos)
        sheet.cell(row=prova, column=3, value=erros)
        sheet.cell(row=prova, column=4, value=pontuacao)

        # Atualize o Excel após pressionar qualquer tecla (exceto 'q')
        workbook.save("pontuacao.xlsx")

    video.release()
    cv2.destroyAllWindows()

@app.route('/video_feed')
def video_feed():
    return Response(processar_video(), mimetype='multipart/x-mixed-replace; boundary=frame')

if __name__ == '__main__':
    app.run(debug=True)
